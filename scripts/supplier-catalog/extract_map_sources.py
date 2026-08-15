#!/usr/bin/env python3
"""Extract official MAP (Minimum Advertised Price) data from the source
manufacturer/distributor price sheets under docs/reference/data/TSW/ into a
single normalized CSV.

Sources:
  - "US MAP MAR26.pdf"            Columbia Tools MAP, effective 2026-03-09
  - "TSW MAPS US HT June 26.pdf"  TSW/Columbia MAP, updated 2026-06-11
  - "2025 Mapp Pricing - TapeTech - EIFS - Decorative - Asgard - FINAL 092425.xlsx"
    TapeTech/EIFS/Decorative/Asgard MAPP, dated 2025-09-24 (sheets: ATF, EIFS,
    Decorative, Asgard)

Output columns: source_file, source_sku, description, map_price
"""
from __future__ import annotations

import argparse
import csv
import re
import tempfile
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
import pdfplumber

HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[1]
TSW_DIR = ROOT / "docs" / "reference" / "data" / "TSW"
DEFAULT_OUTPUT = HERE / "results" / "map" / "map-price-sources.csv"

ITEM_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9.\-/]*$")
PRICE_RE = re.compile(r"^[\d,]+\.\d\d$")


class ExtractionError(RuntimeError):
    pass


def parse_price(raw: str, *, source_sku: str, source: str) -> str:
    try:
        price = Decimal(raw.replace(",", ""))
    except InvalidOperation as exc:
        raise ExtractionError(f"{source}: {source_sku}: invalid MAP price {raw!r}") from exc
    if not price.is_finite() or price <= 0:
        raise ExtractionError(f"{source}: {source_sku}: MAP price must be positive")
    return format(price, ".2f")


def extract_pdf(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines: dict[int, list[dict]] = {}
            for word in page.extract_words():
                lines.setdefault(round(word["top"]), []).append(word)
            for top in sorted(lines):
                words = sorted(lines[top], key=lambda w: w["x0"])
                item, price, desc_tokens = None, None, []
                for word in words:
                    x0, text = word["x0"], word["text"]
                    if x0 < 110 and item is None and ITEM_RE.match(text):
                        item = text
                    elif PRICE_RE.match(text):
                        price = text
                    elif 110 <= x0 < 430:
                        desc_tokens.append(text)
                if item and price:
                    rows.append(
                        {
                            "source_file": path.name,
                            "source_sku": item,
                            "description": " ".join(desc_tokens).strip(),
                            "map_price": parse_price(price, source_sku=item, source=path.name),
                        }
                    )
    if not rows:
        raise ExtractionError(f"{path}: no MAP rows extracted")
    return rows


def extract_xlsx(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    workbook = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_seen = False
        for row in sheet.iter_rows(values_only=True):
            if not header_seen:
                if any(str(v).strip() == "Model" for v in row if v is not None):
                    header_seen = True
                continue
            model, description, price = (row + (None, None, None))[:3]
            if model is None or price is None:
                continue
            model_text = str(model).strip()
            if not model_text or model_text.lower() == "model":
                continue
            rows.append(
                {
                    "source_file": f"{path.name}::{sheet_name}",
                    "source_sku": model_text,
                    "description": str(description).strip() if description else "",
                    "map_price": parse_price(str(price), source_sku=model_text, source=f"{path.name}::{sheet_name}"),
                }
            )
    if not rows:
        raise ExtractionError(f"{path}: no MAP rows extracted")
    return rows


def write_csv_atomic(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["source_file", "source_sku", "description", "map_price"]
    handle = tempfile.NamedTemporaryFile(
        "w", encoding="utf-8-sig", newline="", delete=False, dir=path.parent, prefix=path.name + ".", suffix=".tmp"
    )
    temp_path = Path(handle.name)
    try:
        with handle:
            writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\r\n", extrasaction="raise")
            writer.writeheader()
            writer.writerows(sorted(rows, key=lambda r: (r["source_file"], r["source_sku"])))
        import os

        os.replace(temp_path, path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    rows: list[dict[str, str]] = []
    rows += extract_pdf(TSW_DIR / "US MAP MAR26.pdf")
    rows += extract_pdf(TSW_DIR / "TSW MAPS US HT June 26.pdf")
    rows += extract_xlsx(TSW_DIR / "2025 Mapp Pricing - TapeTech - EIFS - Decorative - Asgard - FINAL 092425.xlsx")

    write_csv_atomic(args.output.resolve(), rows)
    print(f"Extracted {len(rows)} MAP price rows from 3 source files -> {args.output}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ExtractionError as exc:
        print(f"ERROR: {exc}")
        raise SystemExit(1)
